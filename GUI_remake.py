import gi, os, csv, time
import numpy as np
import openpyxl, shutil, datetime, re
from threading import Thread  # Correct import for threading
os.environ["LIBGL_ALWAYS_SOFTWARE"] = "1"
os.environ["GDK_RENDERING"] = "cairo"  # or "gl"
gi.require_version("Gtk", "4.0")
from gi.repository import Gtk, Gio, Gdk  # Import Gdk for applying the CSS
from PIL import Image, ImageDraw
from gi.repository import GdkPixbuf, GLib
from typing import List

def clean_coordinate(value):
    """
    Cleans a coordinate string by ensuring correct decimal placement and preventing unnecessary float rounding.
    """
    if isinstance(value, (int, float)):
        return value  # If it's already a valid number, return as-is

    if isinstance(value, str):
        value = value.strip().replace(" ", "")

        # Ensure it follows the correct format (detect negative and decimal)
        match = re.search(r"-?\d+\.\d+", value)
        if match:
            return float(match.group())  # Convert to float while preserving precision

    return None

def read_coordinates(file_path: str) -> List[float | int]:
    """
    Reads latitude and longitude coordinates from a CSV or Excel (.xlsx) file.
    Supports formats with BREAK lines and standard telemetry headers.
    """
    coordinates = []

    try:
        if file_path.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                if len(row) < 3 or row[1] is None or row[2] is None:
                    continue
                latitude = clean_coordinate(row[1])
                longitude = clean_coordinate(row[2])
                if latitude is not None and longitude is not None:
                    if -90 <= latitude <= 90 and -180 <= longitude <= 180:
                        coordinates.append((latitude, longitude))

        elif file_path.lower().endswith(".csv"):
            with open(file_path, mode='r', newline='', encoding='utf-8') as file:
                reader = csv.reader(file)
                headers = []
                for row in reader:
                    if not row or row[0].startswith("BREAK"):
                        continue

                    # Update headers if it's a header row
                    if row[0].strip() == "Time" and "Latitude" in row and "Longitude" in row:
                        headers = row
                        continue

                    if len(headers) >= 3:
                        try:
                            lat_idx = headers.index("Latitude")
                            lon_idx = headers.index("Longitude")

                            lat_value = row[lat_idx]
                            lon_value = row[lon_idx]

                            latitude = clean_coordinate(lat_value)
                            longitude = clean_coordinate(lon_value)

                            if latitude is not None and longitude is not None:
                                if -90 <= latitude <= 90 and -180 <= longitude <= 180:
                                    coordinates.append((latitude, longitude))
                        except Exception as e:
                            print(f"Skipping row: {row} due to error: {e}")
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")

    return coordinates




class MyWindow(Gtk.Window):
    
    def refresh_image(self, image_path=None, pixbuf=None):
        """
        Refreshes the displayed map image.
        """
        try:
            if pixbuf:
                texture = Gdk.Texture.new_for_pixbuf(pixbuf)
                self.map_image_widget.set_paintable(texture)
            elif image_path:
                file = Gio.File.new_for_path(image_path)
                self.map_image_widget.set_file(file)
            else:
                # Ensure we default to the base image
                base_image_path = "/home/dfec/Desktop/GUI CAPSTONE/Test2Map.png"
                file = Gio.File.new_for_path(base_image_path)
                self.map_image_widget.set_file(file)

            self.content_area.queue_draw()
        except Exception as e:
            print(f"Error refreshing image: {e}")



    def __init__(self, app, rogue_csv_file, discovery_csv_file):
        super().__init__(title="C-UAS Interface 2025")
        self.set_default_size(1000, 600)
        self.set_application(app)  # Link the window to the application

        # Store CSV file paths
        self.rogue_csv_file_path = rogue_csv_file
        self.discovery_csv_file_path = discovery_csv_file

        self.auto_reload =True
        
        #### C H A N G E S ####
        # Store coordinates
        self.rogue_coordinates = []
        self.discovery_coordinates = []

        # Initialize counters for save operations
        self.rogue_save_counter = 1
        self.discovery_save_counter = 1

        # Keep track of already plotted points
        self.plotted_points = set()  # Use a set to store (latitude, longitude) tuples for quick lookup

        def create_legend_item(icon_path, text):
            # Create a horizontal box for the icon and label
            item_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=5)
            # Create the icon image
            icon = Gtk.Image.new_from_file(icon_path)  # Replace with path to your icon image
            icon.set_size_request(50, 50)  # Set the icon size
            # Create the label
            label = Gtk.Label(label=text, halign=Gtk.Align.START)
            # Add the icon and label to the horizontal box
            item_box.append(icon)
            item_box.append(label)
            return item_box
        
        # Create the main container (a horizontal box) for the window content
        main_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=0)
        main_box.set_hexpand(True)  # Allow horizontal expansion
        main_box.set_vexpand(True)  # Allow vertical expansion
        self.set_child(main_box)  # Set main_box as the main widget of the window

        # ------------------------ Left Side Panel -----------------------
        left_panel = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0)
        left_panel.set_size_request(300, -1)  # Set a fixed width for the side panel
        left_panel.add_css_class("side-panel")  # Add a CSS class for styling
        left_panel.set_vexpand(True)  # Allow the side panel to expand vertically
        
        # Create a frame (rectangle) for the title and make it expand horizontally
        legend_frame = Gtk.Frame()
        legend_frame.add_css_class("title-frame")  # Add CSS class for further styling
        legend_label = Gtk.Label(label="Legend", halign=Gtk.Align.CENTER)
        legend_frame.set_child(legend_label)  # Add the label inside the frame
        left_panel.append(legend_frame)
        legend_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        legend_box.set_margin_top(10)

        # Add Items to the Legend
        legend_box.append(create_legend_item("/home/dfec/Desktop/GUI CAPSTONE/drone_reboot.png", "Booting Drone"))
        legend_box.append(create_legend_item("/home/dfec/Desktop/GUI CAPSTONE/DiscoveryDrone_Transparent.png", "Discovery Drone"))
        legend_box.append(create_legend_item("/home/dfec/Desktop/GUI CAPSTONE/RogueDrone_Transparent.png","Rogue Drone"))
        legend_box.append(create_legend_item("/home/dfec/Desktop/GUI CAPSTONE/FortemRadar.png","Radar"))
        legend_box.append(create_legend_item("/home/dfec/Desktop/GUI CAPSTONE/GCS.png","Ground Station"))
        left_panel.append(legend_box)

        # Create a frame (rectangle) for the title and make it expand horizontally
        status_frame = Gtk.Frame()
        status_frame.add_css_class("title-frame")  # Add CSS class for further styling
        status_label = Gtk.Label(label="Status Panel", halign=Gtk.Align.CENTER)
        status_frame.set_child(status_label)  # Add the label inside the frame
        left_panel.append(status_frame)

        # Add the side panel to the main container
        main_box.append(left_panel)

        # ----------------------- Main Content Area (Center) ----------------
        self.content_area = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0)
        self.content_area.add_css_class("content-area")  # Add a CSS class for styling
        self.content_area.set_hexpand(True)  # Allow the content area to expand horizontally
        self.content_area.set_vexpand(True)  # Allow the content area to expand vertically
        main_box.append(self.content_area)

        # Inside MyWindow.__init__ before self.refresh_image(...)
        self.replay_map_image = None  # Holds persistent replay map drawing
        self.map_image_widget = Gtk.Picture()
        self.map_image_widget.set_hexpand(True)
        self.map_image_widget.set_vexpand(True)
        self.content_area.append(self.map_image_widget)

        # Load initial static image map
        self.refresh_image("/home/dfec/Desktop/GUI CAPSTONE/GUI CAPSTONE/Test2Map.png")

        # ----------------------- Right Side Panel ----------------------------
        right_panel = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0)
        right_panel.set_size_request(300, -1)  # Set a fixed width for the side panel
        right_panel.add_css_class("side-panel")  # Add a CSS class for styling
        right_panel.set_vexpand(True)  # Allow the side panel to expand vertically

        # Create a frame (rectangle) for the title and make it expand horizontally
        title_frame = Gtk.Frame()
        title_frame.add_css_class("title-frame")  # Add CSS class for further styling
        right_panel_title = Gtk.Label(label="Controls", halign=Gtk.Align.CENTER)
        title_frame.set_child(right_panel_title)  # Add the label inside the frame
        right_panel.append(title_frame)

        # Create buttons and add them to the right panel ######
        replay_button = Gtk.Button(label="Replay Points")
        replay_button.set_margin_top(10)
        replay_button.set_margin_bottom(10)
        replay_button.set_margin_start(10)
        replay_button.set_margin_end(10)
        replay_button.set_size_request(200,60)
        replay_button.add_css_class("large-button")
        replay_button.connect("clicked", self.on_replay_button_clicked)  # Connect replay function
        right_panel.append(replay_button)

        # Save RogueCoord Btn 
        save_rogue_button = Gtk.Button(label="Save Rogue Coordinates")
        save_rogue_button.connect("clicked", self.on_save_rogue_coords_clicked)
        save_rogue_button.set_margin_top(10)
        save_rogue_button.set_margin_bottom(10)
        save_rogue_button.set_margin_start(10)
        save_rogue_button.set_margin_end(10)
        save_rogue_button.set_size_request(200,60)
        save_rogue_button.add_css_class("save-button")
        right_panel.append(save_rogue_button)

        # Save DiscoveryCoord Btn
        save_discovery_button = Gtk.Button(label="Save Discovery Coordinates")
        save_discovery_button.connect("clicked", self.on_save_discovery_coords_clicked)
        save_discovery_button.set_margin_top(10)
        save_discovery_button.set_margin_top(10)
        save_discovery_button.set_margin_bottom(10)
        save_discovery_button.set_margin_start(10)
        save_discovery_button.set_margin_end(10)
        save_discovery_button.set_size_request(200,60)
        save_discovery_button.add_css_class("save-button")
        right_panel.append(save_discovery_button)

        clear_map_button = Gtk.Button(label="Clear Map")
        clear_map_button.set_margin_top(10)
        clear_map_button.set_margin_top(10)
        clear_map_button.set_margin_top(10)
        clear_map_button.set_margin_bottom(10)
        clear_map_button.set_margin_start(10)
        clear_map_button.set_margin_end(10)
        clear_map_button.set_size_request(200,60)
        clear_map_button.add_css_class("save-button")
        clear_map_button.connect("clicked", self.on_clear_map_clicked)  # Connect to a method
        right_panel.append(clear_map_button)

        reload_data_button = Gtk.Button(label="Reload Data")
        reload_data_button.set_margin_top(10)
        reload_data_button.set_margin_top(10)
        reload_data_button.set_margin_top(10)
        reload_data_button.set_margin_top(10)
        reload_data_button.set_margin_bottom(10)
        reload_data_button.set_margin_start(10)
        reload_data_button.set_margin_end(10)
        reload_data_button.set_size_request(200,60)
        reload_data_button.connect("clicked", self.on_reload_data_clicked)  # Connect to method
        right_panel.append(reload_data_button)

        main_box.append(right_panel)

        # Apply the custom CSS styles
        self.apply_css()

    def apply_css(self):
        """
        Applies custom CSS styling to the window and its widgets.
        """
        try:
            css_provider = Gtk.CssProvider()
            css = b"""
            .side-panel {
                border: 2px solid black;
                padding: 10px;
            }
            .content-area {
                border: 2px solid blue;
                padding: 10px;
            }
            .title-frame {
                background-color: #a9a9a9;
                border: 2px solid black;
                padding: 5px;
            }
            label {
                font-weight: bold;
            }
            .large-button {
                padding: 20px;
                font-size: 18px;
            }
            .save-button {
                padding: 20px;
                font-size: 18px;
            }
            """
            css_provider.load_from_data(css)
            display = Gdk.Display.get_default()
            Gtk.StyleContext.add_provider_for_display(display, css_provider, Gtk.STYLE_PROVIDER_PRIORITY_APPLICATION)
        except Exception as e:
            print(f"Error applying CSS: {e}")

    def save_coordinates_to_excel(self, coordinates, file_name):
        """
        Saves coordinates to an Excel file in the same format as the original file.
        :param coordinates: List of (latitude, longitude) tuples.
        :param file_name: Path to the output Excel file.
        """
        print(f"Saving coordinates: {coordinates} to {file_name}")

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Coordinates"

        # Write the header in a single cell
        sheet.append(["Latitude,Longitude"])

        # Write coordinates to the Excel sheet
        for coord in coordinates:
            # Combine latitude and longitude into one string with a comma, replacing '.' with ','
            combined = f"{coord[0]:.8f},{coord[1]:.8f}".replace('.', ',')
            sheet.append([combined])  # Write the combined string into one cell

        # Save the workbook
        workbook.save(file_name)
        print(f"Saved coordinates to {file_name}")

    def on_clear_map_clicked(self, button):
        """
        Clears the map by resetting it to the base image.
        """
        print("Clear Map button clicked.")
        self.clear_map()
        self.auto_reload = False #Disable auto-reloading if needed

    def on_reload_map_clicked(self, button):
        """
        Resets the map to its original image.
        """
        print("Reload Map button clicked.")
        self.refresh_image("/home/dfec/Desktop/GUI CAPSTONE/Test2Map.png")

    def on_reload_data_clicked(self, button):
        """
        Re-enables auto-reload and updates the map with current data.
        """
        print("Reload Data button clicked. Enabling auto-reload.")
        self.auto_reload = True

        # Force a data update immediately
        self.rogue_coordinates = read_coordinates(self.rogue_csv_file_path)
        self.discovery_coordinates = read_coordinates(self.discovery_csv_file_path)

        self.update_map(self.rogue_coordinates, self.discovery_coordinates)

    def on_save_rogue_coords_clicked(self, button):
        """
        Copies the original RogueCoords.csv file and saves it with a unique timestamp.
        """
        try:
            base_path = "/home/dfec/Desktop/GUI CAPSTONE"
            original_file = os.path.join(base_path, "RogueCoords.csv")  # Ensure it's a CSV file

            # Create a unique filename using a timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file = os.path.join(base_path, f"RogueCoords_Copy_{timestamp}.csv")

            # Copy the file
            shutil.copy(original_file, new_file)

            print(f"Saved a copy of RogueCoords.csv as: {new_file}")

        except Exception as e:
            print(f"Error copying RogueCoords.csv: {e}")

    def on_save_discovery_coords_clicked(self, button):
        """
        Copies the original DiscoveryCoords.xlsx file and saves it with a unique name.
        """
        try:
            base_path = "/home/dfec/Desktop/GUI CAPSTONE"
            original_file = os.path.join(base_path, "DiscoveryCoords.csv")  # Ensure correct file format
            new_file = os.path.join(base_path, f"discovery_coords_copy_{self.discovery_save_counter}.csv")

            shutil.copy(original_file, new_file)  # Copy to retain original structure

            print(f"Saved discovery coordinates as a copy: {new_file}")

            self.discovery_save_counter += 1  # Increment counter
        except Exception as e:
            print(f"Error saving discovery coordinates: {e}")


    def on_replay_button_clicked(self, button):
        """
        Displays a dropdown containing all replay save files (both Rogue & Discovery).
        The user selects a file, and the replay starts automatically.
        """
        # Create a modal dialog
        dialog = Gtk.Dialog(
            transient_for=self,
            modal=True,
            title="Select a Replay Save File",
        )
        dialog.set_default_size(300, 150)

        # Create a dropdown menu
        file_dropdown = Gtk.ComboBoxText()

        # Combine all available replay files (Rogue + Discovery)
        all_replay_files = [
             "RogueCoords_Downsampled_Every5.csv"
        ]

        # Populate the dropdown
        for file in all_replay_files:
            file_dropdown.append_text(file)

        # Select the first file by default
        file_dropdown.set_active(0)

        # Layout for the dialog
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        box.set_margin_top(20)
        box.set_margin_bottom(20)
        box.set_margin_start(20)
        box.set_margin_end(20)
        box.append(file_dropdown)

        # Add buttons to the dialog
        dialog_content_area = dialog.get_content_area()
        dialog_content_area.append(box)
        dialog.add_buttons(
            "OK", Gtk.ResponseType.OK,
            "Cancel", Gtk.ResponseType.CANCEL
        )

        # Show the dialog and handle the response
        dialog.show()

        def on_response(dialog, response):
            if response == Gtk.ResponseType.OK:
                selected_file = file_dropdown.get_active_text()
                if selected_file:
                    print(f"Replay selected for file: {selected_file}")
                    self.replay_points(selected_file, "replay_drone", (1,0,0))  # Start replay
                else:
                    print("No file selected.")
            else:
                print("Replay canceled.")
            dialog.destroy()

        # Connect the response signal
        dialog.connect("response", on_response)

    def clear_map(self):
        """
        Clears the map and resets it to the base image.
        """
        try:
            base_image_path = "/home/dfec/Desktop/GUI CAPSTONE/Test2Map.png"

            print("Resetting map to original base image.")

            # Ensure the image is reloaded correctly
            self.refresh_image(base_image_path)

            # Force UI update
            self.content_area.queue_draw()

        except Exception as e:
            print(f"Error clearing map: {e}")


    def on_clear_map_clicked(self, button):
        """
        Clears the map and resets it to the base image.
        """
        print("Clear Map button clicked.")
        self.clear_map()
        self.auto_reload = False  # Disable auto-reloading if needed


    def start_csv_monitoring(self):
        """
        Starts a background thread to monitor the CSV files for updates.
        """
        def monitor_csv_file():
            while True:
                try:
                    if not self.auto_reload:  # Ensure auto-reload is OFF before processing updates
                        print("Auto-reload disabled. Waiting...")
                        time.sleep(2)  # Sleep longer when disabled to avoid spamming logs
                        continue  # Skip the rest of the loop

                    # Read the rogue and discovery coordinates
                    rogue_coordinates = read_coordinates(self.rogue_csv_file_path)
                    discovery_coordinates = read_coordinates(self.discovery_csv_file_path)

                    # Safely update the map on the main thread
                    GLib.idle_add(self.update_map, rogue_coordinates, discovery_coordinates)

                    # Save the data to class variables for later use
                    self.rogue_coordinates = rogue_coordinates
                    self.discovery_coordinates = discovery_coordinates

                except Exception as e:
                    print(f"Error in monitor_csv_file: {e}")

                time.sleep(0.5)  # Sleep for 500ms before checking again


        # ✅ Passes function without self issue
        monitor_thread = Thread(target=monitor_csv_file)
        monitor_thread.daemon = True  # Ensures the thread stops when the application closes
        monitor_thread.start()


    def update_map(self, rogue_coordinates, discovery_coordinates):
        """
        Updates the map with rogue and discovery coordinates.
        :param rogue_coordinates: List of tuples (latitude, longitude) for rogue drones.
        :param discovery_coordinates: List of tuples (latitude, longitude) for discovery drones.
        """
        try:
            # Clear the map and start with the base image
            if not hasattr(self, "base_map_pixbuf"):
                self.base_map_pixbuf = GdkPixbuf.Pixbuf.new_from_file(
                   "/home/dfec/Desktop/GUI CAPSTONE/Test2Map.png"
                )
            
            # Convert Pixbuf to a PIL Image
            width, height = self.base_map_pixbuf.get_width(), self.base_map_pixbuf.get_height()
            rowstride, n_channels = self.base_map_pixbuf.get_rowstride(), self.base_map_pixbuf.get_n_channels()
            pixels = self.base_map_pixbuf.get_pixels()

            array = np.frombuffer(pixels, dtype=np.uint8).reshape((height, rowstride))

            if n_channels == 1:
                # Convert grayscale to RGB
                image_data = np.stack((array[:, :width],) * 3, axis=-1)
            elif n_channels == 3:
                image_data = array[:, :width * 3].reshape((height, width, 3))
            elif n_channels == 4:
                image_data = array[:, :width * 4].reshape((height, width, 4))[:, :, :3]
            else:
                raise ValueError(f"Unexpected number of channels: {n_channels}")

            pil_image = Image.fromarray(image_data, "RGB")
            draw = ImageDraw.Draw(pil_image)

            # Define map corners (lat/lon)
            lat1, lon1 = 39.019045, -104.894301  # Top-left corner
            lat3, lon3 = 39.017430, -104.894301  # Bottom-left corner
            lon2 = -104.892113  # Top-right corner

            # Calculate lat/lon per pixel
            lat_range = lat1 - lat3
            lon_range = lon2 - lon1
            lat_per_pixel = lat_range / height
            lon_per_pixel = lon_range / width

            # Function to plot points
            def plot_point(coordinate, color):
                latitude, longitude = coordinate
                if not (lat3 <= latitude <= lat1) or not (lon1 <= longitude <= lon2):
                    print(f"Warning: Latitude {latitude} or Longitude {longitude} out of bounds.")
                    return None
                pixel_x = int((longitude - lon1) / lon_per_pixel)
                pixel_y = int((lat1 - latitude) / lat_per_pixel)
                draw.ellipse((pixel_x - 5, pixel_y - 5, pixel_x + 5, pixel_y + 5), fill=color, outline="black")

            # Plot rogue and discovery coordinates
            for coord in rogue_coordinates:
                plot_point(coord, "red")
            for coord in discovery_coordinates:
                plot_point(coord, "green")

            # Convert back to Pixbuf for GTK
            updated_array = np.array(pil_image)
            updated_pixbuf = GdkPixbuf.Pixbuf.new_from_data(
                updated_array.tobytes(),
                GdkPixbuf.Colorspace.RGB,
                False,
                8,
                updated_array.shape[1],
                updated_array.shape[0],
                updated_array.shape[1] * 3,
            )
            texture = Gdk.Texture.new_for_pixbuf(updated_pixbuf)

            # Update the Gtk.Picture widget
            self.map_image_widget.set_paintable(texture)
            self.content_area.queue_draw()
        except Exception as e:
            print(f"Error updating map: {e}")
    
    def replay_points(self, file_name, drone_name, color):
        """
        Replays saved flight path from a selected file.
        """
        try:
            print(f"Replay initiated for {drone_name} drone with file: {file_name}")
            self.auto_reload = False  
            print("Auto-reload permanently disabled during replay.")

            # Load coordinates
            coordinates = read_coordinates(file_name)
            if not coordinates:
                print(f"No coordinates found in {file_name}")
                return

            print(f"Loaded {len(coordinates)} points from {file_name}")

            # Load a fresh copy of the base map as the replay canvas
            self.replay_map_image = Image.open("/home/dfec/Desktop/GUI CAPSTONE/Test2Map.png").convert("RGB")
            draw = ImageDraw.Draw(self.replay_map_image)

            # Store coordinates and initialize index
            self.replay_coordinates = coordinates
            self.replay_index = 0

            def plot_next_point():
                if self.replay_index < len(self.replay_coordinates):
                    lat, lon = self.replay_coordinates[self.replay_index]
                    print(f"Plotting {drone_name} point {self.replay_index + 1}/{len(self.replay_coordinates)}: ({lat}, {lon})")

                    # Convert to pixels
                    pixel_x, pixel_y = self.convert_to_pixels(lat, lon)

                    # Draw point
                    draw.ellipse((pixel_x - 5, pixel_y - 5, pixel_x + 5, pixel_y + 5), fill="red", outline="black")

                    # Draw connecting path
                    if self.replay_index > 0:
                        prev_lat, prev_lon = self.replay_coordinates[self.replay_index - 1]
                        prev_x, prev_y = self.convert_to_pixels(prev_lat, prev_lon)
                        draw.line([(prev_x, prev_y), (pixel_x, pixel_y)], fill="red", width=2)

                    # Update map widget with current replay frame
                    updated_array = np.array(self.replay_map_image)
                    updated_pixbuf = GdkPixbuf.Pixbuf.new_from_data(
                        updated_array.tobytes(),
                        GdkPixbuf.Colorspace.RGB,
                        False,
                        8,
                        updated_array.shape[1],
                        updated_array.shape[0],
                        updated_array.shape[1] * 3,
                    )
                    texture = Gdk.Texture.new_for_pixbuf(updated_pixbuf)
                    self.map_image_widget.set_paintable(texture)
                    self.content_area.queue_draw()

                    self.replay_index += 1
                    return True  # Keep replaying
                else:
                    print(f"{drone_name.capitalize()} Drone Replay completed.")
                    return False  # Stop replay

            # Start the timed animation
            GLib.timeout_add(200, plot_next_point)

        except Exception as e:
            print(f"Error in replay_points: {e}")



    def convert_to_pixels(self, lat, lon):
        """
        Converts latitude and longitude to pixel coordinates on the map.
        Ensures correct scaling so points are mapped accurately.
        """
        # Define reference points for the top-left, bottom-left, and top-right corners
        lat1, lon1 = 39.019045, -104.894301  # Top-left
        lat3, lon3 = 39.017430, -104.894301  # Bottom-left
        lon2 = -104.892113  # Top-right

        # Get map dimensions (match your actual base image size)
        width, height = 1000, 600  # Adjust these values to match the base image

        # Calculate the range of latitude and longitude
        lat_range = lat1 - lat3
        lon_range = lon2 - lon1

        # Convert lat/lon to pixel coordinates
        try:
            pixel_x = int(((lon - lon1) / lon_range) * width)
            pixel_y = int(((lat1 - lat) / lat_range) * height)

            print(f"Lat: {lat}, Lon: {lon} -> Pixel: ({pixel_x}, {pixel_y})")  # Debugging output
            return pixel_x, pixel_y
        except Exception as e:
            print(f"Error in convert_to_pixels: {e}")
            return 0, 0  # Return (0,0) if an error occurs
    
    def update_map_safe(self, rogue_coordinates, discovery_coordinates):
        GLib.idle_add(self.update_map, rogue_coordinates, discovery_coordinates)

class MyApp(Gtk.Application):
    def __init__(self, rogue_csv_file, discovery_csv_file):
        super().__init__(application_id="org.example.myapp", flags=Gio.ApplicationFlags.FLAGS_NONE)
        self.rogue_csv_file = rogue_csv_file
        self.discovery_csv_file = discovery_csv_file   

    def do_activate(self):
        # Define the path to the CSV file
        win = MyWindow(self, self.rogue_csv_file, self.discovery_csv_file)
        win.present()
        win.start_csv_monitoring()

def main():

    discovery_csv_file = "/home/dfec/cuas_24-25/agent_core/disco_position.csv"
    rogue_csv_file = "/home/dfec/cuas_24-25/agent_core/rogue_position.csv"
    #rogue_csv_file = "/home/dfec/Desktop/GUI CAPSTONE/RogueCoords.csv"
    #discovery_csv_file = "/home/dfec/Desktop/GUI CAPSTONE/DiscoveryCoords.csv"
    app = MyApp(rogue_csv_file,discovery_csv_file)
    app.run(None)

if __name__ == "__main__":
    main()
